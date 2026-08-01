"""Lossless, bounded readers for roster-shaped CSV and Excel files.

The existing student importer intentionally turns a table into dictionaries as
quickly as possible.  That remains the right behavior for known templates, but
an interactive column-mapping screen needs the structure that dictionaries
discard: column positions, duplicate headers, and unmodified cell values.  The
types in this module form that smaller, lossless boundary.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any, BinaryIO, Literal, Sequence, TextIO

from seattrellis.io.json_files import InputFileError
from seattrellis.optional import MissingOptionalDependencyError


DEFAULT_MAX_ROSTER_FILE_BYTES = 20 * 1024 * 1024
DEFAULT_MAX_ROSTER_ROWS = 10_000
DEFAULT_MAX_ROSTER_COLUMNS = 256


@dataclass(frozen=True)
class RosterTableLimits:
    """Resource limits applied before and while parsing an uploaded roster."""

    max_file_bytes: int = DEFAULT_MAX_ROSTER_FILE_BYTES
    max_rows: int = DEFAULT_MAX_ROSTER_ROWS
    max_columns: int = DEFAULT_MAX_ROSTER_COLUMNS

    def __post_init__(self) -> None:
        for name, value in (
            ("max_file_bytes", self.max_file_bytes),
            ("max_rows", self.max_rows),
            ("max_columns", self.max_columns),
        ):
            if isinstance(value, bool) or not isinstance(value, int) or value < 1:
                raise ValueError(f"{name} must be a positive integer")


DEFAULT_ROSTER_TABLE_LIMITS = RosterTableLimits()


@dataclass(frozen=True)
class RosterColumn:
    """One physical column, identified by position rather than header text."""

    index: int
    raw_header: Any

    @property
    def header(self) -> str:
        """Return a displayable header without changing ``raw_header``."""

        if self.raw_header is None:
            return ""
        return str(self.raw_header).strip()


@dataclass(frozen=True)
class RosterRow:
    """One physical data row with its source row number and raw cell values."""

    row_number: int
    cells: tuple[Any, ...]

    def cell(self, column_index: int) -> Any:
        """Return a cell by physical column, or ``None`` for a short row."""

        if isinstance(column_index, bool) or not isinstance(column_index, int):
            raise TypeError("column_index must be an integer")
        if column_index < 0:
            raise IndexError("column_index cannot be negative")
        if column_index >= len(self.cells):
            return None
        return self.cells[column_index]


@dataclass(frozen=True)
class RosterTable:
    """A lossless roster table suitable for preview and column mapping."""

    columns: tuple[RosterColumn, ...]
    rows: tuple[RosterRow, ...]
    source_format: Literal["csv", "xlsx"]
    sheet_name: str | None = None
    headerless: bool = False

    def __post_init__(self) -> None:
        if not self.columns:
            raise ValueError("A roster table must contain a header row")
        expected_indices = tuple(range(len(self.columns)))
        actual_indices = tuple(column.index for column in self.columns)
        if actual_indices != expected_indices:
            raise ValueError("Roster columns must have contiguous zero-based indices")

    @property
    def column_count(self) -> int:
        return len(self.columns)

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def headers(self) -> tuple[str, ...]:
        """Display headers in physical order, including duplicates and blanks."""

        return tuple(column.header for column in self.columns)


def read_roster_table(
    path: str | Path,
    *,
    limits: RosterTableLimits = DEFAULT_ROSTER_TABLE_LIMITS,
) -> RosterTable:
    """Read a CSV/XLSX roster without collapsing duplicate headers.

    Files are size-checked before opening. CSV input is decoded as UTF-8 with
    an optional BOM. Excel workbooks use openpyxl's read-only mode and are
    closed in a ``finally`` block, including when a limit or parsing error is
    raised.
    """

    source = Path(path)
    if not source.exists():
        raise InputFileError(f"Roster file not found: {source}")
    if not source.is_file():
        raise InputFileError(f"Roster input is not a file: {source}")
    try:
        size = source.stat().st_size
    except OSError as exc:
        raise InputFileError(f"Could not inspect roster file {source}: {exc}") from exc
    _check_file_size(size, limits)

    suffix = source.suffix.lower()
    try:
        if suffix == ".csv":
            with source.open("r", encoding="utf-8-sig", newline="") as stream:
                return _read_csv_stream(stream, limits=limits)
        if suffix in {".xlsx", ".xlsm"}:
            with source.open("rb") as stream:
                return _read_xlsx_stream(stream, limits=limits)
        if suffix == ".xls":
            raise InputFileError(
                f"Legacy .xls files are not supported for {source}; save as .xlsx or CSV."
            )
        raise InputFileError(
            f"Unsupported roster file format for {source}: {source.suffix or '<none>'}"
        )
    except (InputFileError, MissingOptionalDependencyError):
        raise
    except UnicodeDecodeError as exc:
        raise InputFileError(f"Roster CSV must be UTF-8 text: {source}: {exc}") from exc
    except (csv.Error, OSError) as exc:
        raise InputFileError(f"Could not read roster file {source}: {exc}") from exc


def read_roster_table_bytes(
    data: bytes,
    *,
    filename: str,
    limits: RosterTableLimits = DEFAULT_ROSTER_TABLE_LIMITS,
) -> RosterTable:
    """Read an in-memory upload while applying the same limits as file input."""

    if not isinstance(data, bytes):
        raise TypeError("data must be bytes")
    if not isinstance(filename, str) or not filename.strip():
        raise ValueError("filename must be a non-empty string")
    _check_file_size(len(data), limits)
    suffix = Path(filename).suffix.lower()
    try:
        if suffix == ".csv":
            text = io.TextIOWrapper(io.BytesIO(data), encoding="utf-8-sig", newline="")
            try:
                return _read_csv_stream(text, limits=limits)
            finally:
                text.close()
        if suffix in {".xlsx", ".xlsm"}:
            return _read_xlsx_stream(io.BytesIO(data), limits=limits)
        if suffix == ".xls":
            raise InputFileError(
                "Legacy .xls files are not supported; save as .xlsx or CSV."
            )
        raise InputFileError(
            f"Unsupported roster file format: {Path(filename).suffix or '<none>'}"
        )
    except (InputFileError, MissingOptionalDependencyError):
        raise
    except UnicodeDecodeError as exc:
        raise InputFileError(f"Roster CSV must be UTF-8 text: {exc}") from exc
    except (csv.Error, OSError) as exc:
        raise InputFileError(f"Could not read roster upload: {exc}") from exc


def _read_csv_stream(
    stream: TextIO,
    *,
    limits: RosterTableLimits,
) -> RosterTable:
    reader = csv.reader(stream, strict=True)
    try:
        raw_headers = next(reader)
    except StopIteration as exc:
        raise InputFileError("Roster CSV is empty; a header row is required.") from exc
    _check_column_count(len(raw_headers), limits)
    if not raw_headers:
        raise InputFileError("Roster CSV has no columns in its header row.")

    raw_rows: list[tuple[int, tuple[Any, ...]]] = []
    for row_number, values in enumerate(reader, start=2):
        if len(raw_rows) >= limits.max_rows:
            raise InputFileError(
                f"Roster has more than the allowed {limits.max_rows} data rows."
            )
        _check_column_count(len(values), limits)
        raw_rows.append((row_number, tuple(values)))

    return _build_roster_table(
        raw_headers,
        raw_rows,
        source_format="csv",
        limits=limits,
    )


def _read_xlsx_stream(
    stream: BinaryIO,
    *,
    limits: RosterTableLimits,
) -> RosterTable:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise MissingOptionalDependencyError("Excel roster preview", "excel") from exc

    workbook = None
    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
        sheet = workbook.active
        # Dimension metadata lets us reject accidentally formatted million-row
        # sheets before iterating them. The streaming checks below remain the
        # authority because third-party generators can publish stale metadata.
        if sheet.max_column > limits.max_columns:
            _check_column_count(sheet.max_column, limits)
        if sheet.max_row > limits.max_rows + 1:
            raise InputFileError(
                f"Roster has more than the allowed {limits.max_rows} data rows."
            )

        iterator = sheet.iter_rows(values_only=True)
        try:
            raw_headers = tuple(next(iterator))
        except StopIteration as exc:
            raise InputFileError(
                "Roster workbook is empty; a header row is required."
            ) from exc
        _check_column_count(len(raw_headers), limits)
        if not raw_headers:
            raise InputFileError("Roster workbook has no columns in its header row.")

        raw_rows: list[tuple[int, tuple[Any, ...]]] = []
        for row_number, values in enumerate(iterator, start=2):
            if len(raw_rows) >= limits.max_rows:
                raise InputFileError(
                    f"Roster has more than the allowed {limits.max_rows} data rows."
                )
            _check_column_count(len(values), limits)
            raw_rows.append((row_number, tuple(values)))

        return _build_roster_table(
            raw_headers,
            raw_rows,
            source_format="xlsx",
            limits=limits,
            sheet_name=sheet.title,
        )
    except (InputFileError, MissingOptionalDependencyError):
        raise
    except Exception as exc:
        # openpyxl exposes several parser-specific exceptions. Keep those
        # implementation details behind the established input error boundary.
        raise InputFileError(f"Could not read roster workbook: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _check_file_size(size: int, limits: RosterTableLimits) -> None:
    if size > limits.max_file_bytes:
        raise InputFileError(
            f"Roster file is {size} bytes; the limit is {limits.max_file_bytes} bytes."
        )


def _check_column_count(count: int, limits: RosterTableLimits) -> None:
    if count > limits.max_columns:
        raise InputFileError(
            f"Roster has {count} columns; the limit is {limits.max_columns}."
        )


_HEADER_HINTS = frozenset(
    {
        "id",
        "sid",
        "student",
        "studentid",
        "studentnumber",
        "studentname",
        "name",
        "fullname",
        "phone",
        "mobile",
        "phonenumber",
        "gender",
        "sex",
        "height",
        "heightcm",
        "score",
        "grade",
        "vision",
        "needs",
        "notes",
        "tags",
        "姓名",
        "学生姓名",
        "学号",
        "学生编号",
        "编号",
        "电话",
        "手机号",
        "性别",
        "身高",
        "成绩",
        "总分",
        "视力",
        "特殊需求",
        "备注",
        "标签",
    }
)


def _build_roster_table(
    raw_headers: Sequence[Any],
    raw_rows: Sequence[tuple[int, tuple[Any, ...]]],
    *,
    source_format: Literal["csv", "xlsx"],
    limits: RosterTableLimits,
    sheet_name: str | None = None,
) -> RosterTable:
    """Build a table and retain a first row that is actually roster data."""

    headers = tuple(raw_headers)
    headerless = bool(raw_rows and _looks_like_headerless(headers, raw_rows[0][1]))
    data_rows = list(raw_rows)
    if headerless:
        data_rows.insert(0, (1, tuple(headers)))
        headers = tuple(f"Column {index + 1}" for index in range(len(headers)))
        if len(data_rows) > limits.max_rows:
            raise InputFileError(
                f"Roster has more than the allowed {limits.max_rows} data rows."
            )

    rows: list[RosterRow] = []
    for row_number, values in data_rows:
        _check_column_count(len(values), limits)
        if len(values) > len(headers):
            raise InputFileError(
                f"Roster row {row_number} has {len(values)} cells but the header has "
                f"only {len(headers)} columns."
            )
        rows.append(RosterRow(row_number=row_number, cells=tuple(values)))

    columns = tuple(
        RosterColumn(index=index, raw_header=value)
        for index, value in enumerate(headers)
    )
    return RosterTable(
        columns=columns,
        rows=tuple(rows),
        source_format=source_format,
        sheet_name=sheet_name,
        headerless=headerless,
    )


def _looks_like_headerless(
    headers: Sequence[Any],
    first_data_row: Sequence[Any],
) -> bool:
    """Use conservative value-shape checks to catch common headerless exports."""

    normalized = {
        "".join(
            character
            for character in str(value).strip().casefold()
            if character.isalnum()
        )
        for value in headers
        if str(value).strip()
    }
    if normalized & _HEADER_HINTS:
        return False
    return _looks_like_record(headers) and _looks_like_record(first_data_row)


def _looks_like_record(values: Sequence[Any]) -> bool:
    for value in values:
        text = str(value).strip()
        if not text:
            continue
        if text.isdigit() and len(text) >= 4:
            return True
        if any("\u4e00" <= character <= "\u9fff" for character in text):
            return True
    return False


__all__ = [
    "DEFAULT_MAX_ROSTER_COLUMNS",
    "DEFAULT_MAX_ROSTER_FILE_BYTES",
    "DEFAULT_MAX_ROSTER_ROWS",
    "DEFAULT_ROSTER_TABLE_LIMITS",
    "RosterColumn",
    "RosterRow",
    "RosterTable",
    "RosterTableLimits",
    "read_roster_table",
    "read_roster_table_bytes",
]
