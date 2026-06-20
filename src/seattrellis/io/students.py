from __future__ import annotations

import csv
from collections.abc import Iterable, Mapping
from math import isnan
from pathlib import Path
from typing import Any

try:
    from pydantic.v1 import ValidationError
except ImportError:  # pragma: no cover - pydantic v1.
    from pydantic import ValidationError
from openpyxl import load_workbook

from seattrellis.io.json_files import InputFileError
from seattrellis.models.student import Student

COLUMN_ALIASES = {
    "student_id": {"student_id", "id", "sid", "学号", "学生编号", "编号"},
    "name": {"name", "姓名", "学生姓名"},
    "gender": {"gender", "sex", "性别"},
    "height_cm": {"height_cm", "height", "身高", "身高cm", "身高_cm"},
    "score": {"score", "成绩", "总分", "分数"},
    "vision": {"vision", "视力"},
    "notes": {"notes", "note", "备注", "说明"},
    "tags": {"tags", "tag", "标签"},
    "needs": {"needs", "need", "特殊需求", "需求"},
}


def read_students(path: str | Path) -> list[Student]:
    source = Path(path)
    if not source.exists():
        raise InputFileError(f"Student file not found: {source}")
    try:
        if source.suffix.lower() == ".csv":
            rows = _read_csv_rows(source)
        elif source.suffix.lower() in {".xlsx", ".xlsm"}:
            rows = _read_excel_rows(source)
        elif source.suffix.lower() == ".xls":
            raise InputFileError(f"Legacy .xls files are not supported for {source}; save as .xlsx or CSV.")
        else:
            raise InputFileError(f"Unsupported student file format for {source}: {source.suffix}")
    except InputFileError:
        raise
    except Exception as exc:
        raise InputFileError(f"Could not read student file {source}: {exc}") from exc
    try:
        return students_from_records(rows)
    except ValueError as exc:
        raise InputFileError(f"Invalid student file {source}: {exc}") from exc


def students_from_dataframe(frame: Any) -> list[Student]:
    """Build students from a pandas-like DataFrame without requiring pandas at import time."""

    if hasattr(frame, "to_dict"):
        return students_from_records(frame.to_dict(orient="records"))
    raise TypeError("students_from_dataframe expects a pandas-like DataFrame.")


def students_from_records(records: Iterable[Mapping[str, Any]]) -> list[Student]:
    rows = list(records)
    columns = [str(column) for row in rows for column in row.keys()]
    column_map = _build_column_map(columns)
    students: list[Student] = []
    for row_index, row in enumerate(rows):
        data: dict[str, Any] = {}
        attributes: dict[str, Any] = {}
        for column, raw_value in row.items():
            value = _clean_value(raw_value)
            if value is None:
                continue
            canonical = column_map.get(str(column))
            if canonical:
                data[canonical] = value
            else:
                attributes[str(column)] = value
        if attributes:
            data["attributes"] = attributes
        if not data:
            continue
        try:
            students.append(Student(**data))
        except ValidationError as exc:
            messages = "; ".join(error["msg"] for error in exc.errors())
            raise ValueError(f"row {row_index + 2}: {messages}") from exc
    if not students:
        raise ValueError("No valid students found.")
    _validate_unique_keys(students)
    return students


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def _read_excel_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    records: list[dict[str, Any]] = []
    for values in rows[1:]:
        records.append({headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]})
    return records


def _build_column_map(columns: Iterable[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    normalized_aliases = {
        alias.lower().replace(" ", "").replace("_", ""): canonical
        for canonical, aliases in COLUMN_ALIASES.items()
        for alias in aliases
    }
    for column in columns:
        text = str(column)
        normalized = text.lower().replace(" ", "").replace("_", "")
        if normalized in normalized_aliases:
            mapping[text] = normalized_aliases[normalized]
    return mapping


def _clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and isnan(value):
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return value


def _validate_unique_keys(students: list[Student]) -> None:
    student_ids = [student.student_id for student in students if student.student_id]
    duplicate_ids = sorted({student_id for student_id in student_ids if student_ids.count(student_id) > 1})
    if duplicate_ids:
        raise ValueError(f"Duplicate student_id values: {', '.join(duplicate_ids)}")

    keys = [student.key for student in students]
    duplicates = sorted({key for key in keys if keys.count(key) > 1})
    if duplicates:
        raise ValueError(f"Duplicate student identifiers: {', '.join(duplicates)}")
