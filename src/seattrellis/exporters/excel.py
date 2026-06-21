from __future__ import annotations

from pathlib import Path

from seattrellis.models.snapshot import SeatingSnapshot
from seattrellis.optional import MissingOptionalDependencyError


def export_excel(snapshot: SeatingSnapshot, output: str | Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise MissingOptionalDependencyError("Excel export", "excel") from exc

    path = Path(output)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Seating"
    _write_grid(sheet, snapshot, Alignment, Font, PatternFill, get_column_letter)

    assignments = workbook.create_sheet("Assignments")
    assignments.append(["student_key", "student_name", "seat_id"])
    for assignment in snapshot.assignments:
        assignments.append([assignment.student_key, assignment.student_name, assignment.seat_id])

    workbook.save(path)
    return path


def _write_grid(sheet, snapshot: SeatingSnapshot, Alignment, Font, PatternFill, get_column_letter) -> None:
    min_row, max_row, min_col, max_col = _bounds(snapshot)
    assignment_by_seat = {assignment.seat_id: assignment for assignment in snapshot.assignments}
    seat_by_position = {(seat.row, seat.col): seat for seat in snapshot.layout.seats}

    title = f"{snapshot.layout.name} - {snapshot.created_at:%Y-%m-%d %H:%M}"
    sheet.cell(row=1, column=1, value=title)
    sheet.cell(row=1, column=1).font = Font(bold=True)

    disabled_fill = PatternFill("solid", fgColor="D9D9D9")
    enabled_fill = PatternFill("solid", fgColor="EAF4FF")
    fixed_fill = PatternFill("solid", fgColor="DFF2BF")

    for row in range(min_row, max_row + 1):
        excel_row = row - min_row + 3
        sheet.row_dimensions[excel_row].height = 42
        for col in range(min_col, max_col + 1):
            excel_col = col - min_col + 1
            cell = sheet.cell(row=excel_row, column=excel_col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            seat = seat_by_position.get((row, col))
            if seat is None:
                cell.value = ""
                cell.fill = disabled_fill
            elif not seat.enabled:
                cell.value = f"{seat.seat_id}\n--"
                cell.fill = disabled_fill
            else:
                assignment = assignment_by_seat.get(seat.seat_id)
                cell.value = f"{seat.seat_id}\n{assignment.student_name if assignment else ''}"
                cell.fill = fixed_fill if assignment else enabled_fill
            sheet.column_dimensions[get_column_letter(excel_col)].width = 18


def _bounds(snapshot: SeatingSnapshot) -> tuple[int, int, int, int]:
    rows = [seat.row for seat in snapshot.layout.seats]
    cols = [seat.col for seat in snapshot.layout.seats]
    return min(rows), max(rows), min(cols), max(cols)
